import csv
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


ALIASES = {
    "certNo": ["证书编号", "证书号", "certno", "certificate", "编号"],
    "factoryNo": ["出厂编号", "出厂号", "factoryno", "出厂编码"],
    "sendUnit": ["送检单位", "使用单位", "单位名称", "企业名称", "companyname"],
    "instrumentName": ["仪表名称", "压力表名称", "名称", "instrumentname"],
    "modelSpec": ["型号规格", "规格型号", "型号", "modelspec"],
    "manufacturer": ["制造单位", "制造厂家", "生产厂家", "manufacturer"],
    "verificationStd": ["检定依据", "检定规程", "校准依据", "verificationstd"],
    "conclusion": ["检定结论", "结论", "conclusion"],
    "verificationDate": ["检定日期", "校验日期", "检验日期", "verificationdate"],
    "gaugeStatus": ["压力表状态", "状态", "gaugestatus"],
    "installLocation": ["安装位置", "安装地点", "使用位置", "location"],
    "equipmentName": ["所属设备", "设备名称", "设备", "equipmentname"],
    "district": ["辖区", "区县", "district"],
}


def normalize_header(value):
    return str(value or "").strip().replace(" ", "").replace("_", "").lower()


def build_header_mapping(headers):
    mapped = {}
    normalized_headers = [normalize_header(item) for item in headers]
    for canonical, alias_list in ALIASES.items():
        alias_set = {normalize_header(alias) for alias in alias_list}
        for index, header in enumerate(normalized_headers):
            if header in alias_set:
                mapped[canonical] = index
                break
    return mapped


def row_to_payload(row, mapping):
    payload = {}
    for key, index in mapping.items():
        if index < len(row):
            value = row[index]
            payload[key] = "" if value is None else str(value).strip()
    return payload


def load_csv_rows(file_path):
    last_error = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            with open(file_path, "r", encoding=encoding, newline="") as fp:
                return list(csv.reader(fp))
        except Exception as exc:
            last_error = exc
    raise last_error or RuntimeError("无法读取 CSV 文件")


def load_excel_rows(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    rows = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
    return rows


def compact_rows(rows):
    return [row for row in rows if any(str(cell or "").strip() for cell in row)]


def parse_file(file_path):
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = load_csv_rows(file_path)
    else:
        rows = load_excel_rows(file_path)

    rows = compact_rows(rows)
    if not rows:
      return {"rows": []}

    headers = rows[0]
    mapping = build_header_mapping(headers)
    data_rows = []
    for offset, row in enumerate(rows[1:], start=2):
        payload = row_to_payload(row, mapping)
        if not any(payload.values()):
            continue
        payload["rowNumber"] = offset
        data_rows.append(payload)

    return {"rows": data_rows}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"rows": []}, ensure_ascii=False))
        sys.exit(0)

    result = parse_file(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False))
